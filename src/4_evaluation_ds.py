"""
Lucas Correia
LIACS | Leiden University
Einsteinweg 55 | 2333 CC Leiden | The Netherlands
"""

import os

os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import numpy as np
import tensorflow as tf
from tqdm.contrib import itertools
from dotenv import dotenv_values
from sklearn import metrics
from utilities import detection_class
import pandas as pd
import openpyxl
from dtaidistance import dtw

seeds = [1, 2, 3]
folds = [0, 1, 2]
splits = ['1day', '1week', '2weeks', '3weeks', '4weeks']

# Declare constants
BUDGET = 10
MISLABEL_PROB = 0

# Load variables in .env file
config = dotenv_values("../.env")
data_path = config['data_path']
model_path = config['model_path']

results = []
for seed, fold_idx in itertools.product(seeds, folds):
    # Set fixed seed for random operations
    np.random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)

    query_list = []
    query_score_list = []
    for split_idx, split in enumerate(splits):
        data_load_path = os.path.join(data_path, '2_preprocessed', 'fold_' + str(fold_idx))

        model_name = 'tevae_' + split + '_' + str(fold_idx) + '_1'  # Fixed model seed due to focus on query strategy
        model_load_path = os.path.join(model_path, model_name)

        # Load tf validation data to extract window size
        tfdata_val = tf.data.Dataset.load(os.path.join(data_load_path, split, 'val'))

        detector = detection_class.AnomalyDetector(
            model_path=model_load_path,
            window_size=tfdata_val.element_spec.shape[0],
            sampling_rate=2,
            original_sampling_rate=10,
            calculate_delay=True,
            label_keyword='normal',
            mislabel_prob=MISLABEL_PROB,
        )

        # Load input data
        train_list = detector.load_pickle(os.path.join(data_load_path, split, 'train.pkl'))
        val_list = detector.load_pickle(os.path.join(data_load_path, split, 'val.pkl'))
        test_list = detector.load_pickle(os.path.join(data_load_path, split, 'test.pkl'))

        # Load inference data
        train_detection_score_list = detector.load_pickle(os.path.join(model_load_path, 'train_detection_score.pkl'))
        val_detection_score_list = detector.load_pickle(os.path.join(model_load_path, 'val_detection_score.pkl'))
        test_detection_score_list = detector.load_pickle(os.path.join(model_load_path, 'test_detection_score.pkl'))

        # Combine training and validation data to get candidate list
        candidate_list = train_list + val_list
        candidate_score_list = train_detection_score_list + val_detection_score_list

        # Remove previously queried data from candidate list
        queried_filenames_list = [query_ts.dtype.metadata['file_name'] for query_ts in query_list]
        candidate_list = [candidate_ts for candidate_ts in candidate_list if candidate_ts.dtype.metadata['file_name'] not in queried_filenames_list]
        candidate_score_list = [candidate_score for candidate_score in candidate_score_list if candidate_score.dtype.metadata['file_name'] not in queried_filenames_list]

        # Create look-up table for DTW distances between specific time series
        precomputed_distances = {}


        def compute_dtw_distance(ts1, ts2):
            key = tuple(sorted((ts1.dtype.metadata['file_name'], ts2.dtype.metadata['file_name'])))
            # If distance for time series pair is not in look-up table, compute and store it
            if key not in precomputed_distances:
                precomputed_distances[key] = dtw.distance_fast(
                    ts1.astype(np.dtype(np.double)),
                    ts2.astype(np.dtype(np.double))
                )
            return precomputed_distances[key]

        # Query strategy
        for _ in range(BUDGET):
            if len(query_score_list) == 0:
                # Random selection for the first budget iteration in '1day' split
                selection_idx = np.random.randint(0, high=len(candidate_score_list))
            else:
                # Compute distance matrix between query and candidates
                query_distances = np.array([
                    [compute_dtw_distance(query_score, candidate_score) for candidate_score in candidate_score_list]
                    for query_score in query_score_list
                ])
                # Find the most similar candidate to any query
                min_distance_idx = np.argmin(query_distances.mean(axis=0))
                sim_candidate = candidate_score_list[min_distance_idx]
                # Compute distance of all candidates to the most similar candidate
                sim_candidate_distances = np.array([
                    compute_dtw_distance(sim_candidate, candidate_score) for candidate_score in candidate_score_list
                ])
                # Select the candidate most dissimilar to the most similar candidate
                selection_idx = np.argmax(sim_candidate_distances)
            # Update query lists and remove selected candidate
            query_list.append(candidate_list[selection_idx])
            query_score_list.append(candidate_score_list[selection_idx])
            del candidate_list[selection_idx]
            del candidate_score_list[selection_idx]

        # Extract groundtruth labels from query_list and test_list
        query_groundtruth_list = detector.extract_groundtruth(query_list)
        test_groundtruth_list = detector.extract_groundtruth(test_list)

        # Flip labels randomly with a given probability
        query_contaminated_list = detector.corrupt_labels(query_groundtruth_list)

        # Grid search for thresholds using labelled data
        f1_list = []
        reduced_query_score_list = np.concatenate(query_score_list).ravel()
        percentile_array = np.arange(0, 100.01, 0.01)
        for threshold_percentile in percentile_array:
            threshold_temp = np.percentile(reduced_query_score_list, threshold_percentile)
            groundtruth_labels_temp, predicted_labels_temp, _ = detector.evaluate_online(
                detection_score_list=query_score_list,
                groundtruth_list=query_contaminated_list,
                threshold=threshold_temp,
            )
            f1_list.append(metrics.f1_score(groundtruth_labels_temp, predicted_labels_temp, zero_division=0.0))
        f1_list = np.vstack(f1_list)
        threshold = np.percentile(reduced_query_score_list, percentile_array[np.argmax(f1_list)]).astype(float)

        # Evaluate using threshold
        groundtruth_labels, predicted_labels, total_delays = detector.evaluate_online(
            detection_score_list=test_detection_score_list,
            groundtruth_list=test_groundtruth_list,
            threshold=threshold,
        )

        results.append({
            'Seed': seed,
            'Fold': fold_idx,
            'Split': split,
            'F1': metrics.f1_score(groundtruth_labels, predicted_labels, zero_division=0.0),
            'Precision': metrics.precision_score(groundtruth_labels, predicted_labels, zero_division=0.0),
            'Recall': metrics.recall_score(groundtruth_labels, predicted_labels, zero_division=0.0),
            'Delay': np.mean(total_delays),
            'Threshold': threshold
        })

results = pd.DataFrame(results)

if not os.path.isfile(os.path.join(model_path, 'results.xlsx')):
    # Create and save a valid Excel file
    wb = openpyxl.Workbook()
    wb.save(os.path.join(model_path, 'results.xlsx'))

# Use a try-finally block to ensure proper handling
try:
    with pd.ExcelWriter(os.path.join(model_path, 'results.xlsx'), mode='a', if_sheet_exists='overlay') as writer:
        results.to_excel(
            writer,
            index=False,
            sheet_name=f'ds_{BUDGET}_{int(MISLABEL_PROB * 100)}'
        )
finally:
    # Cleanup: Remove default 'Sheet' if it exists
    try:
        workbook = openpyxl.load_workbook(os.path.join(model_path, 'results.xlsx'))
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        workbook.save(os.path.join(model_path, 'results.xlsx'))
    except Exception as e:
        print(f"Error cleaning up sheets: {e}")
