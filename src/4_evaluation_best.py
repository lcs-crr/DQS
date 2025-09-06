"""
Lucas Correia
LIACS | Leiden University
Einsteinweg 55 | 2333 CC Leiden | The Netherlands
"""

import os

os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
import numpy as np
import tensorflow as tf
from tqdm.contrib import itertools
from dotenv import dotenv_values
from sklearn import metrics
from utilities import detection_class
import pandas as pd
import openpyxl

folds = [0, 1, 2]
splits = ['1day', '1week', '2weeks', '3weeks', '4weeks']

# Load variables in .env file
config = dotenv_values("../.env")
data_path = os.path.join(config['data_path'], 'dsq')
model_path = os.path.join(config['model_path'], 'dsq')

results = []
for fold_idx, split in itertools.product(folds, splits):
    data_load_path = os.path.join(data_path, '2_preprocessed', 'fold_' + str(fold_idx))

    model_name = 'tevae_' + split + '_' + str(fold_idx) + '_1'  # Fixed model seed due to focus on query strategy
    model_load_path = os.path.join(model_path, model_name)

    # Load tf validation data to extract window size
    tfdata_val = tf.data.Dataset.load(os.path.join(data_load_path, split, 'val'))

    detector = detection_class.AnomalyDetector(
        model_path=model_load_path,
        window_size=tfdata_val.element_spec.shape[0],
        sampling_rate=2,
        original_sampling_rate=10,
        calculate_delay=True,
        label_keyword='normal',
    )

    # Load test data
    test_list = detector.load_pickle(os.path.join(data_load_path, split, 'test.pkl'))

    # Load test inference results
    test_detection_score_list = detector.load_pickle(os.path.join(model_load_path, 'test_detection_score.pkl'))

    # Extract groundtruth labels from test_list
    test_groundtruth_list = detector.extract_groundtruth(test_list)

    f1_list = []
    reduced_test_detection_score = np.concatenate(test_detection_score_list).ravel()
    percentile_array = np.arange(0, 100.01, 0.01)
    for threshold_percentile in percentile_array:
        threshold_temp = np.percentile(reduced_test_detection_score, threshold_percentile)
        groundtruth_labels_temp, predicted_labels_temp, _ = detector.evaluate_online(
            detection_score_list=test_detection_score_list,
            groundtruth_list=test_groundtruth_list,
            threshold=threshold_temp,
        )
        f1_list.append(metrics.f1_score(groundtruth_labels_temp, predicted_labels_temp, zero_division=0.0))
    f1_list = np.vstack(f1_list)
    threshold = np.percentile(reduced_test_detection_score, percentile_array[np.argmax(f1_list)]).astype(float)

    groundtruth_labels, predicted_labels, total_delays = detector.evaluate_online(
        detection_score_list=test_detection_score_list,
        groundtruth_list=test_groundtruth_list,
        threshold=threshold,
    )

    results.append({
        'Seed': 'x',
        'Fold': fold_idx,
        'Split': split,
        'F1': metrics.f1_score(groundtruth_labels, predicted_labels, zero_division=0.0),
        'Precision': metrics.precision_score(groundtruth_labels, predicted_labels, zero_division=0.0),
        'Recall': metrics.recall_score(groundtruth_labels, predicted_labels, zero_division=0.0),
        'Delay': np.mean(total_delays),
        'Threshold': threshold
    })

results = pd.DataFrame(results)

if not os.path.isfile(os.path.join(model_path, 'results.xlsx')):
    # Create and save a valid Excel file
    wb = openpyxl.Workbook()
    wb.save(os.path.join(model_path, 'results.xlsx'))

# Use a try-finally block to ensure proper handling
try:
    with pd.ExcelWriter(os.path.join(model_path, 'results.xlsx'), mode='a', if_sheet_exists='overlay') as writer:
        results.to_excel(
            writer,
            index=False,
            sheet_name=f'best'
        )
finally:
    # Cleanup: Remove default 'Sheet' if it exists
    try:
        workbook = openpyxl.load_workbook(os.path.join(model_path, 'results.xlsx'))
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        workbook.save(os.path.join(model_path, 'results.xlsx'))
    except Exception as e:
        print(f"Error cleaning up sheets: {e}")
